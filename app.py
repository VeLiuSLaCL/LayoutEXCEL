import io
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


st.set_page_config(page_title="Acomodador de Layout", layout="wide")
st.title("Acomodador de archivos al layout base")

st.markdown(
    """
Sube primero el **layout vigente** y después el **listado de archivos de finiquitos / nómina**.

La app:
- toma el encabezado del layout como estructura principal,
- acomoda cada archivo aunque no tenga el mismo orden de columnas,
- deja en blanco las columnas faltantes,
- agrega al final las columnas extra que no existan en el layout,
- coloca **Archivo origen** al **final final**,
- resalta las columnas extra con otro color,
- deja la **fila 2 en blanco** y comienza los datos desde la **fila 3**,
- respeta los importes numéricos con formato **0.00**.
"""
)


# =========================
# Utilidades
# =========================
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")         # azul oscuro
HEADER_FONT = Font(color="FFFFFF", bold=True)
EXTRA_HEADER_FILL = PatternFill("solid", fgColor="C65911")   # naranja
EXTRA_DATA_FILL = PatternFill("solid", fgColor="FCE4D6")     # naranja claro
SOURCE_HEADER_FILL = PatternFill("solid", fgColor="5B9BD5")  # azul medio
SOURCE_DATA_FILL = PatternFill("solid", fgColor="DDEBF7")    # azul claro
BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

MONEY_NUMBER_FORMAT = "0.00"


def clean_header(value):
    if value is None:
        return ""
    value = str(value).replace("\n", " ").replace("\r", " ")
    value = re.sub(r"\s+", " ", value).strip()
    return value


def validate_extension(uploaded_file):
    ext = Path(uploaded_file.name).suffix.lower()
    return ext in SUPPORTED_EXTENSIONS


def is_effectively_empty_row(row):
    for value in row:
        if value not in (None, ""):
            return False
    return True


def normalize_decimal_value(value):
    if value is None or isinstance(value, bool):
        return None

    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"))

    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(Decimal("0.01"))

    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None

        # Conserva textos comunes que no son importes
        if re.search(r"[A-Za-z]", text):
            return None

        # Limpieza básica para números con separadores
        cleaned = text.replace(",", "")
        try:
            return Decimal(cleaned).quantize(Decimal("0.01"))
        except InvalidOperation:
            return None

    return None


def apply_numeric_money_format(cell):
    decimal_value = normalize_decimal_value(cell.value)
    if decimal_value is not None:
        cell.value = float(decimal_value)
        cell.number_format = MONEY_NUMBER_FORMAT


@st.cache_data(show_spinner=False)
def read_workbook_headers_and_rows(file_bytes: bytes, filename: str):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return {
            "filename": filename,
            "sheet_name": ws.title,
            "headers": [],
            "rows": [],
            "max_column": 0,
        }

    headers = [clean_header(v) for v in header_row]
    rows = []
    for row in rows_iter:
        row = list(row)
        if len(row) < len(headers):
            row += [None] * (len(headers) - len(row))
        row = row[:len(headers)]
        if is_effectively_empty_row(row):
            continue
        rows.append(row)

    return {
        "filename": filename,
        "sheet_name": ws.title,
        "headers": headers,
        "rows": rows,
        "max_column": len(headers),
    }


def build_output_workbook(layout_headers, processed_files, include_source_col=True):
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Concentrado"

    normalized_layout = [clean_header(h) for h in layout_headers if clean_header(h)]

    extras_in_order = []
    extras_seen = set()

    for file_data in processed_files:
        for header in file_data["headers"]:
            if not header:
                continue
            if header not in normalized_layout and header not in extras_seen:
                extras_seen.add(header)
                extras_in_order.append(header)

    final_headers = list(normalized_layout)
    final_headers.extend(extras_in_order)
    if include_source_col:
        final_headers.append("Archivo origen")

    # encabezado fila 1
    for col_idx, header in enumerate(final_headers, start=1):
        cell = ws_out.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

        if header == "Archivo origen":
            cell.fill = SOURCE_HEADER_FILL
        elif header in extras_in_order:
            cell.fill = EXTRA_HEADER_FILL
        else:
            cell.fill = HEADER_FILL

    # fila 2 vacía intencionalmente
    output_row = 3

    for file_data in processed_files:
        header_to_index = {h: idx for idx, h in enumerate(file_data["headers"]) if h}

        for source_row in file_data["rows"]:
            output_values = []

            for layout_header in normalized_layout:
                idx = header_to_index.get(layout_header)
                output_values.append(source_row[idx] if idx is not None and idx < len(source_row) else None)

            for extra_header in extras_in_order:
                idx = header_to_index.get(extra_header)
                output_values.append(source_row[idx] if idx is not None and idx < len(source_row) else None)

            if include_source_col:
                output_values.append(file_data["filename"])

            for col_idx, value in enumerate(output_values, start=1):
                cell = ws_out.cell(row=output_row, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = BORDER

                header_name = final_headers[col_idx - 1]
                if header_name == "Archivo origen":
                    cell.fill = SOURCE_DATA_FILL
                elif header_name in extras_in_order:
                    cell.fill = EXTRA_DATA_FILL

                if header_name != "Archivo origen":
                    apply_numeric_money_format(cell)

            output_row += 1

    # congelar encabezado
    ws_out.freeze_panes = "A3"
    ws_out.auto_filter.ref = f"A1:{get_column_letter(ws_out.max_column)}1"
    ws_out.row_dimensions[1].height = 28
    ws_out.row_dimensions[2].height = 8

    # Ajuste de anchos
    max_widths = {}
    sample_limit = min(ws_out.max_row, 300)

    for col_idx in range(1, ws_out.max_column + 1):
        header = clean_header(ws_out.cell(row=1, column=col_idx).value)
        max_len = min(max(len(header), 10), 45)

        for row_idx in range(3, sample_limit + 1):
            value = ws_out.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            val_len = len(str(value))
            if val_len > max_len:
                max_len = min(val_len, 45)

        if header == "Archivo origen":
            width = 24
        elif header in extras_in_order:
            width = max(14, min(max_len + 2, 24))
        else:
            width = max(12, min(max_len + 2, 24))

        max_widths[col_idx] = width

    for col_idx, width in max_widths.items():
        ws_out.column_dimensions[get_column_letter(col_idx)].width = width

    return wb_out, extras_in_order, final_headers


def workbook_to_bytes(workbook):
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# =========================
# UI
# =========================
st.subheader("1) Sube el layout vigente")
layout_file = st.file_uploader(
    "Layout base",
    type=["xlsx", "xlsm", "xltx", "xltm"],
    accept_multiple_files=False,
    key="layout_file",
)

st.subheader("2) Sube los archivos a acomodar")
source_files = st.file_uploader(
    "Archivos de entrada",
    type=["xlsx", "xlsm", "xltx", "xltm"],
    accept_multiple_files=True,
    key="source_files",
)

include_source_col = st.checkbox(
    "Agregar columna 'Archivo origen' al resultado",
    value=True,
    help="Se colocará al final final del archivo para que sea la última columna.",
)

if source_files:
    st.info(f"Archivos cargados: **{len(source_files)}**")

if layout_file and not validate_extension(layout_file):
    st.error("El layout no tiene una extensión soportada.")
    st.stop()

invalid_sources = []
if source_files:
    invalid_sources = [f.name for f in source_files if not validate_extension(f)]

if invalid_sources:
    st.error("Estos archivos no tienen una extensión soportada: " + ", ".join(invalid_sources))
    st.stop()

if layout_file:
    layout_data = read_workbook_headers_and_rows(layout_file.getvalue(), layout_file.name)
    if not layout_data["headers"]:
        st.error("No pude leer el encabezado del layout.")
        st.stop()

    layout_headers = [h for h in layout_data["headers"] if h]

    with st.expander("Ver columnas del layout"):
        st.write(f"Total de columnas del layout: **{len(layout_headers)}**")
        st.dataframe({"Columna": layout_headers}, use_container_width=True)

if layout_file and source_files:
    processed_files = []
    resumen = []

    for f in source_files:
        data = read_workbook_headers_and_rows(f.getvalue(), f.name)
        processed_files.append(data)

        extras = [h for h in data["headers"] if h and h not in layout_headers]
        missing = [h for h in layout_headers if h not in data["headers"]]

        resumen.append({
            "Archivo": data["filename"],
            "Hoja": data["sheet_name"],
            "Columnas detectadas": len([h for h in data["headers"] if h]),
            "Registros detectados": len(data["rows"]),
            "Columnas extra vs layout": len(extras),
            "Columnas faltantes vs layout": len(missing),
        })

    st.subheader("Resumen de archivos")
    st.dataframe(resumen, use_container_width=True)

    if st.button("Generar archivo acomodado", type="primary"):
        with st.spinner("Procesando archivos..."):
            wb_out, extras_in_order, final_headers = build_output_workbook(
                layout_headers=layout_headers,
                processed_files=processed_files,
                include_source_col=include_source_col,
            )
            output_bytes = workbook_to_bytes(wb_out)

        total_rows = max(wb_out.active.max_row - 2, 0)

        st.success("Archivo generado correctamente.")
        st.write(f"Registros consolidados: **{total_rows}**")
        st.write(f"Columnas finales: **{len(final_headers)}**")
        st.write(f"Columnas extra agregadas al final: **{len(extras_in_order)}**")

        if extras_in_order:
            with st.expander("Ver columnas extra agregadas al final"):
                st.dataframe({"Columna extra": extras_in_order}, use_container_width=True)

        st.download_button(
            label="Descargar resultado",
            data=output_bytes,
            file_name="resultado_layout_acomodado.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        with st.expander("Vista previa de columnas finales"):
            st.dataframe({"Columna final": final_headers}, use_container_width=True)
